from fastapi import FastAPI
from pydantic import BaseModel
from openpyxl import load_workbook

# 初始化 FastAPI 应用
app = FastAPI()

# 读取 Excel 文件并构建翻译字典
excel_file = "i18n_RAG.xlsx"
translation_map = {}

wb = load_workbook(excel_file)
ws = wb.active

# 获取表头
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    row_dict = dict(zip(headers, row))
    original = row_dict["zh-TW"].strip()
    translation_map[original] = {
        "zh-TW": str(row_dict.get("zh-TW") or "").strip(),
        "zh-CN": str(row_dict.get("zh-CN") or "").strip(),
        "en-US": str(row_dict.get("en-US") or "").strip(),
        "ja-JP": str(row_dict.get("ja-JP") or "").strip(),
        "ko-KR": str(row_dict.get("ko-KR") or "").strip(),
    }

# 输入结构
class LookupRequest(BaseModel):
    phrases: list


# POST 接口：批量翻译查找
@app.post("/lookup")
async def lookup(request: LookupRequest):
    matched = []
    unmatched = []

    for phrase in request.phrases:
        translation = translation_map.get(phrase)
        if translation:
            matched.append({
                "o": phrase,
                "t": translation
            })
        else:
            unmatched.append({
                "o": phrase
            })

    return {
        "matched": matched,
        "unmatched": unmatched
    }
